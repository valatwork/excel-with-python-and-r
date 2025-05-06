from openpyxl import load_workbook
import pandas as pd

def read_single_sheet(workbook, sheet_name):
    # Select the sheet
    sheet = workbook[sheet_name]

    # Extract the values (including headers)
    sheet_data_raw = sheet.values

    # Separate the headers into a variable
    columns = next(sheet_data_raw)[0:]

    # Create a DataFrame based on the second and subsequent lines of data with the header as column names
    return pd.DataFrame(sheet_data_raw, columns=columns)

def read_multiple_sheets(file_path):

    # Load the Excel file
    workbook = load_workbook(file_path, data_only=True)

    # Get all sheet names
    sheet_names = workbook.sheetnames

    # Cycle through the sheet names, load the data for each, and concatenate them into a single DataFrame
    return pd.concat([read_single_sheet(workbook=workbook, sheet_name=sheet_name) for sheet_name in sheet_names], ignore_index=True)

# Define the file path and sheet names
file_path = './data/iris_data.xlsx'

# Read the data from multiple sheets
consolidated_data = read_multiple_sheets(file_path)

# Display the consolidated data
print(consolidated_data.head())